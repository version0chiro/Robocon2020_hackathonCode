# USAGE
# python recognize_faces_video.py --encodings encodings.pickle
# python recognize_faces_video.py --encodings encodings.pickle --output output/jurassic_park_trailer_output.avi --display 0

# import the necessary packages
import os
from datetime import date

from tensorflow.keras.applications.mobilenet_v2 import preprocess_input
from tensorflow.keras.preprocessing.image import img_to_array
from tensorflow.keras.models import load_model
from imutils.video import VideoStream
import numpy as np
import argparse
import imutils
import time
import cv2
from face_detection import FaceDetection

import os
from imutils.video import VideoStream
import face_recognition
import argparse
import imutils
import pickle
import time
import cv2
import time
import pandas as pd
from collections import Counter
import numpy as np
import time
# from detect_mask_video import detect_and_predict_mask

# writer = cv2.VideoWriter(args["output"], fourcc, 20,
#     (frame.shape[1], frame.shape[0]), True)

Fcount = 500
 #enter the number of frames you wanna save
vs = cv2.VideoCapture(0)
global gMask
global gBox
ret,frame = vs.read()
# video = cv2.VideoCapture("/dev/video1") # check this and uncommetn if you are using a jetson nano

# We need to check if camera
# is opened previously or not
if (vs.isOpened() == False):
    print("Error reading video file")

# We need to set resolutions.
# so, convert them from float to integer.
frame_width = int(vs.get(3))
frame_height = int(vs.get(4))

size = (frame_width, frame_height)

# Below VideoWriter object will create
# a frame of above defined The output
# is stored in 'filename.avi' file.
result = cv2.VideoWriter('videos/filename.avi',
                         cv2.VideoWriter_fourcc(*'MJPG'),
                         10, size)



details=pd.read_excel('details.xlsx',index_col=0)
atte=pd.read_excel('attendance.xlsx',index_col=0)


def spo2Func():
    final_sig = []

    frameCount = 1

    def dethreding(z):

        return signal.detrend(z)

        # T=len(z)
        # lamba=10
        # I  = np.identity(T)
        # filt = [1,-2,1]* np.ones((1,T-2),dtype=np.int).T
        # D2 = scipy.sparse.spdiags(data.T, (range(0,3)),T-2,T)
        # detrended_sig =
        # return z_stat


    def face_detect_and_thresh(frame):
        # lower = np.array([0, 1,0], dtype = "uint8")
        # upper = np.array([179, 255, 255], dtype = "uint8")
        lower = np.array([0, 58, 50] , dtype="uint8")
        upper = np.array([30, 255, 255], dtype="uint8")
        # for detecting skin lower and upper are used for creating a mask
        fd = FaceDetection()
        frame, face_frame, ROI1, ROI2, status, mask, extra, box = fd.face_detect(frame)
        frame = extra
        # cv2.imshow("face0",extra)
        converted = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)
        skinMask = cv2.inRange(converted, lower, upper)
        cv2.imshow("face0", converted)
        cv2.imshow("face3", skinMask)
        # apply a series of erosions and dilations to the mask
        # using an elliptical kernel
        kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (11, 11))
        skinMask = cv2.erode(skinMask, kernel, iterations=2)
        skinMask = cv2.dilate(skinMask, kernel, iterations=2)
        # blur the mask to help remove noise, then apply the
        # mask to the frame
        skinMask = cv2.GaussianBlur(skinMask, (3, 3), 0)
        skin = cv2.bitwise_and(frame, frame, mask=skinMask)
        cv2.imshow("skin2", skin)
        gMask = skinMask
        global gBox
        gBox = box
        # show the skin in the image along with the mask
        # cv2.imshow("images", np.hstack([frame, skin]))
        # cv2.imshow("mask",np.hstack([mask]))

        for _ in range(np.shape(mask)[0]):
            for j in range(np.shape(mask)[1]):
                if mask[_][j] == 1:
                    mask[_][j] = 1
                else:
                    mask[_][j] = 0
        mask = mask.astype(int)
        return skin, frame, mask


    def spartialAverage(thresh, frame):
        a = list(np.argwhere(maskT == 1))
        # x=[i[0] for i in a]
        # y=[i[1] for i in a]
        # p=[x,y]
        ind_img = (np.vstack((a)))
        sig_fin = np.zeros([np.shape(ind_img)[0], 3])
        test_fin = []
        for i in range(np.shape(ind_img)[0]):
            sig_temp = frame[ind_img[i, 0], ind_img[i, 1], :]
            sig_temp = sig_temp.reshape((1, 3))
            if sig_temp.any() != 0:
                sig_fin = np.concatenate((sig_fin, sig_temp))
        # print(sig_fin)
        for _ in sig_fin:
            if sum(_) > 0:
                # test_fin=np.concatenate((test_fin,_))
                test_fin.append(_)
        # print("min=>")
        a = [item for item in sig_fin if sum(item) > 0]
        # print(min(a, key=sum))
        min_value = sum(min(a, key=sum))
        max_value = sum(max(a, key=sum))
        # print(sum1)
        img_rgb_mean = np.nanmean(test_fin, axis=0)
        print(img_rgb_mean)
        return img_rgb_mean, min_value, max_value


    def MeanRGB(thresh, frame, img_rgb, last_stage, min_value, max_value):
        cv2.imshow("threshh", img_rgb)
        print("==<>>")
        # print(img_rgb)
        # cv2.waitKey()
        # print(img_rgb[0])
        # thresh=thresh.reshape((1,3))
        # img_rgb_mean=np.nanmean(thresh,axis=0)
        a = [item for item in img_rgb[0] if (sum(item) < max_value and sum(item) > 200)]
        # a = filter(lambda (x,y,z) : i+j+k>764 ,frame[0])
        # print(a[1:10])
        # img_temp = [item for item in img_rgb if sum(item)>764]
        # print(frame[0])
        # print(img_temp)
        # print(np.mean(a, axis=(0)))
        if a:
            print("==>")
            # print(a)
            print("==>")
            img_mean = np.mean(a, axis=(0))
            # print(img_mean)

            return img_mean[::-1]
        else:
            return last_stage


    def MeanRGB2(thresh, frame, img_rgb, min_value, max_value):
        cv2.imshow("threshh", img_rgb)
        print("==<>>")
        # print(img_rgb)
        # cv2.waitKey()
        # print(img_rgb[0])
        # thresh=thresh.reshape((1,3))
        # img_rgb_mean=np.nanmean(thresh,axis=0)
        a = [item for item in img_rgb[0] if (sum(item) < max_value and sum(item) > min_value)]
        # a = filter(lambda (x,y,z) : i+j+k>764 ,frame[0])
        # print(a[1:10])
        # img_temp = [item for item in img_rgb if sum(item)>764]
        # print(frame[0])
        # print(img_temp)
        # print(np.mean(a, axis=(0)))

        print("==>")
        # print(a)
        print("==>")
        img_mean = np.mean(a, axis=(0))
        # print(img_mean)

        return img_mean[::-1]


    def MeanRGB(thresh, frame, img_rgb, last_stage, max_value, min_value):
        cv2.imshow("threshh", img_rgb)
        print("==<>>")
        # print(img_rgb)
        # cv2.waitKey()
        # print(img_rgb[0])
        # thresh=thresh.reshape((1,3))
        # img_rgb_mean=np.nanmean(thresh,axis=0)
        a = [item for item in img_rgb[0] if (sum(item) < 764 and sum(item) > 200)]
        # a = filter(lambda (x,y,z) : i+j+k>764 ,frame[0])
        # print(a[1:10])
        # img_temp = [item for item in img_rgb if sum(item)>764]
        # print(frame[0])
        # print(img_temp)
        # print(np.mean(a, axis=(0)))
        if a:
            print("==>")
            # print(a)
            print("==>")
            img_mean = np.mean(a, axis=(0))
            # print(img_mean)

            return img_mean[::-1]
        else:
            return last_stage


    def MeanRGB2(thresh, frame, img_rgb):
        cv2.imshow("threshh", img_rgb)
        print("==<>>")
        # print(img_rgb)
        # cv2.waitKey()
        # print(img_rgb[0])
        # thresh=thresh.reshape((1,3))
        # img_rgb_mean=np.nanmean(thresh,axis=0)
        a = [item for item in img_rgb[0] if (sum(item) < 764 and sum(item) > 200)]
        # 764 being the max value of skin in HSV range

        # 200 being the min value of hsv range


        # a = filter(lambda (x,y,z) : i+j+k>764 ,frame[0])
        # print(a[1:10])
        # img_temp = [item for item in img_rgb if sum(item)>764]
        # print(frame[0])
        # print(img_temp)
        # print(np.mean(a, axis=(0)))

        # print("==>")
        # print(a)
        # print("==>")
        img_mean = np.mean(a, axis=(0))
        # print(img_mean)

        return img_mean[::-1]


    def face_detect_and_thresh2(frame):
        lower = np.array([0, 58, 50] , dtype="uint8")
        upper = np.array([30, 255, 255], dtype="uint8")
        face_frame = frame[gBox[0]:gBox[1], gBox[2]:gBox[3]]
        # cv2.imshow("face_testing",face_frame)
        frame = face_frame
        cv2.imshow("testing", frame)
        converted = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)
        skinMask = cv2.inRange(converted, lower, upper)
        # apply a series of erosions and dilations to the mask
        # using an elliptical kernel
        kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (11, 11))
        skinMask = cv2.erode(skinMask, kernel, iterations=2)
        skinMask = cv2.dilate(skinMask, kernel, iterations=2)
        # blur the mask to help remove noise, then apply the
        # mask to the frame
        skinMask = cv2.GaussianBlur(skinMask, (3, 3), 0)
        skin = cv2.bitwise_and(frame, frame, mask=skinMask)
        cv2.imshow("skin", skin)
        # print(np.mean(skin, axis=(0, 1)))
        # cv2.imshow("face_testing",skin)
        return skin, frame


    def preprocess(z1, z2, detrended_RGB, window_size, size_video, duration, frame):
        temp = (int(size_video/duration))
        f = frame-2

        main_R = []
        main_B = []
        out = []
        for i in range(len(detrended_RGB)-f+1):

            temp_R = z1[i:i+f-1]
            temp_B = z2[i:i+f-1]
            p = [list(a) for a in zip(temp_R, temp_B)]

            out.append(p)
            # if not main_R:
            #     main_R.append(temp_R)
            # else:
            #     main_R=[main_R,temp_R]
            #
            # if not main_B:
            #     main_B.append(temp_B)
            # else:
            #     main_B=[main_B,temp_B]

        # out=[main_R,main_B]
        # print(out[0])
        return out[0]


    def SPooEsitmate(final_sig, video_size, frames, seconds):
        A = 101.6
        B = 5.834

        #  A and B are empirically determined coefficients, Iac and Idc are respectively the amplitudes of the pulsatile (ac) and dc components
        # the determination of these values were done on the basis of the paper published
        z1 = [item[0] for item in final_sig]
        z3 = [item[2] for item in final_sig]
        #The mean average value of Blue and Red  is taken from the set of frames
        SPO_pre = []
        for _ in range(len(z1)):
            SPO_pre.append([z1[_], z3[_]])
        Spo2 = preprocess(z1, z3, SPO_pre, 10, video_size, 10, frames)

        R_temp = [item[0] for item in Spo2]
        DC_R_comp = np.mean(R_temp)
        AC_R_comp = np.std(R_temp)

        I_r = AC_R_comp/DC_R_comp

        B_temp = [item[1] for item in Spo2]
        DC_B_comp = np.mean(B_temp)
        AC_B_comp = np.std(B_temp)

        I_b = AC_B_comp/DC_B_comp
        SpO2_value = (A-B*((I_b*650)/(I_r*950)))
        # Valye 650 and 950 was chosen as the wavelengths of red and near infrared
        return SpO2_value


    # main code begins from here and all the above helper fuctions are called


    # Enter the path of video you wanna give into the code
    cap = cv2.VideoCapture("videos/filename.avi")

    # Calculates the total number of frames present in the video
    length = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    fps = cap.get(cv2.CAP_PROP_FPS)      # OpenCV2 version 2 used "CV_CAP_PROP_FPS"
    frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    print(frame_count)
    duration = frame_count/fps
    seconds = duration % 60

    print(length)

    if (cap.isOpened() == False):
        print("Error opening video file")  # Catching anykind of error while opening video

    while(cap.isOpened()):
        ret, frame = cap.read()

        # cv2.waitKey()
        if frameCount == 1:
            # cv2.imshow("testing if flipped",frame)  #just run this if you have issue with inverted frames, can result in fail detection
            # cv2.waitKey()
            # cv2.imshow("testing if flipped2",imutils.rotate_bound(frame, 90))  #just run this if you have issue with inverted frames, can result in fail detection
            # cv2.waitKey()
            # frame= imutils.rotate_bound(frame, 90)
            firstFrame = frame
            # The first frame is sent to the fuction to make a face mask
            thresh, img_rgb, maskT = face_detect_and_thresh(firstFrame)
            cv2.imshow("img_rgb", img_rgb)
            # cv2.waitKey()
            frameCount += 1

        if ret == True:
            # cv2.imshow("testing if flipped",frame)  #just run this if you have issue with inverted frames, can result in fail detection
            frameCount += 1
            # frame= imutils.rotate_bound(frame, 90)

            print(frameCount)
            cv2.imshow('Frame', frame)

            start = time.time()
            # rest of frames are sent to a function to have a threshold/overlap with mask
            thresh, img_rgb = face_detect_and_thresh2(frame)
            end = time.time()
            print(end - start)

            # start = time.time()
            #
            # final_sig.append(spartialAverage(maskT,frame)) # The frames are then sent to spartialAverage to find RGB mean values
            # end = time.time()
            # print(end - start)

            start = time.time()
            if final_sig:
                # The frames are then sent to spartialAverage to find RGB mean values
                final_sig.append(MeanRGB(maskT, frame, img_rgb, final_sig[-1], min_value, max_value))
            else:
                temp, min_value, max_value = spartialAverage(maskT, frame)
                final_sig.append(temp)
            end = time.time()
            print(end - start)

            #  Those rgb values are appened to final_sig which will be used for spo2 estimation
            if cv2.waitKey(25) & 0xFF == ord('q'):  # end the loop if your press 'q' or video reaches end
                break
            # if frameCount==32:
            #     break

      # Break the loop
        else:
            break

    # When everything done, release
    # the video capture object
    cap.release()

    # Closes all the frames
    cv2.destroyAllWindows()


    # For debugging you can try saving the signal as a pickle and run usingPick.py
    pickle.dump(final_sig, open("spo2.p", "wb"))
    print(final_sig)
    # the final signal list is sent to SPooEsitmate function with length of the video
    result = SPooEsitmate(final_sig, length, length, seconds)
    print(result)
    return result

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name,header=False, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

def add_to_list(str_name, atte, details,t,Spo2Count):
	# atte = atte
	# t = time.localtime()
    current_time = time.strftime("%H:%M:%S", t)
    d = details[details['Name'] == str_name]
    d['Time-Stamp'] = current_time
    d['SpO2_value']=Spo2Count
    a = d.index[0]
    return atte.append(d)


def detect_and_predict_mask(frame, faceNet, maskNet):
	# grab the dimensions of the frame and then construct a blob
	# from it
	(h, w) = frame.shape[:2]
	blob = cv2.dnn.blobFromImage(frame, 1.0, (300, 300),
		(104.0, 177.0, 123.0))

	# pass the blob through the network and obtain the face detections
	faceNet.setInput(blob)
	detections = faceNet.forward()

	# initialize our list of faces, their corresponding locations,
	# and the list of predictions from our face mask network
	faces = []
	locs = []
	preds = []

	# loop over the detections
	for i in range(0, detections.shape[2]):
		# extract the confidence (i.e., probability) associated with
		# the detection
		confidence = detections[0, 0, i, 2]

		# filter out weak detections by ensuring the confidence is
		# greater than the minimum confidence
		if confidence > args["confidence"]:
			# compute the (x, y)-coordinates of the bounding box for
			# the object
			box = detections[0, 0, i, 3:7] * np.array([w, h, w, h])
			(startX, startY, endX, endY) = box.astype("int")

			# ensure the bounding boxes fall within the dimensions of
			# the frame
			(startX, startY) = (max(0, startX), max(0, startY))
			(endX, endY) = (min(w - 1, endX), min(h - 1, endY))

			# extract the face ROI, convert it from BGR to RGB channel
			# ordering, resize it to 224x224, and preprocess it
			face = frame[startY:endY, startX:endX]
			face = cv2.cvtColor(face, cv2.COLOR_BGR2RGB)
			face = cv2.resize(face, (224, 224))
			face = img_to_array(face)
			face = preprocess_input(face)

			# add the face and bounding boxes to their respective
			# lists
			faces.append(face)
			locs.append((startX, startY, endX, endY))

	# only make a predictions if at least one face was detected
	if len(faces) > 0:
		# for faster inference we'll make batch predictions on *all*
		# faces at the same time rather than one-by-one predictions
		# in the above `for` loop
		faces = np.array(faces, dtype="float32")
		preds = maskNet.predict(faces, batch_size=32)

	# return a 2-tuple of the face locations and their corresponding
	# locations
	return (locs, preds)

# construct the argument parser and parse the arguments
ap = argparse.ArgumentParser()
ap.add_argument("-e", "--encodings", default="encoding2.pickle",required=True,
	help="path to serialized db of facial encodings")
ap.add_argument("-o", "--output", type=str,
	help="path to output video")
ap.add_argument("-y", "--display", type=int, default=1,
	help="whether or not to display output frame to screen")
ap.add_argument("-d", "--detection-method", type=str, default="hog",
	help="face detection model to use: either `hog` or `cnn`")
ap.add_argument("-f", "--face", type=str,
	default="face_detector",
	help="path to face detector model directory")
ap.add_argument("-m", "--model", type=str,
	default="model/mask_detector.model",
	help="path to trained face mask detector model")
ap.add_argument("-c", "--confidence", type=float, default=0.5,
	help="minimum probability to filter weak detections")
args = vars(ap.parse_args())

# load our serialized face detector model from disk
print("[INFO] loading face detector model...")
prototxtPath = os.path.sep.join([args["face"], "deploy.prototxt"])
weightsPath = os.path.sep.join([args["face"],
	"res10_300x300_ssd_iter_140000.caffemodel"])
faceNet = cv2.dnn.readNet(prototxtPath, weightsPath)

# load the face mask detector model from disk
print("[INFO] loading face mask detector model...")
maskNet = load_model(args["model"])

# initialize the video stream and allow the camera sensor to warm up


# load the known faces and embeddings
print("[INFO] loading encodings...")
data = pickle.loads(open(args["encodings"], "rb").read())

model = load_model(args["model"])

# initialize the video stream and pointer to output video file, then
# allow the camera sensor to warm up
print("[INFO] starting video stream...")
# vs = cv2.VideoCapture(0)
writer = None
time.sleep(2.0)

# loop over frames from the video file stream
while True:
	# grab the frame from the threaded video stream
    ret,frame = vs.read()
    while(Fcount):
        ret, frame = vs.read()

        if ret == True:

    # Write the frame into the
    # file 'filename.avi'
            result.write(frame)
            Fcount=Fcount-1
    # Display the frame
    # saved in the file
            # cv2.imshow('Frame', frame)

    # Press S on keyboard
    # to stop the process


    # Break the loop
        else:
            break

    result.release()

    # convert the input frame from BGR to RGB then resize it to have
	# a width of 750px (to speedup processing)
    rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    rgb = imutils.resize(frame, width=750)
    r = frame.shape[1] / float(rgb.shape[1])

	# detect the (x, y)-coordinates of the bounding boxes
	# corresponding to each face in the input frame, then compute
	# the facial embeddings for each face
    boxes = face_recognition.face_locations(rgb,model=args["detection_method"])
    encodings = face_recognition.face_encodings(rgb, boxes)
    names = []

	# loop over the facial embeddings
    for encoding in encodings:
	    # attempt to match each face in the input image to our known
        # encodings
        matches = face_recognition.compare_faces(data["encodings"],	encoding)
        name = "Unknown"

        # check to see if we have found a match
        if True in matches:
	        # find the indexes of all matched faces then initialize a
            # dictionary to count the total number of times each face
            # was matched
            matchedIdxs = [i for (i, b) in enumerate(matches) if b]
            counts = {}

            # loop over the matched indexes and maintain a count for
            # each recognized face face
            for i in matchedIdxs:
                name = data["names"][i]
                counts[name] = counts.get(name, 0) + 1
            print(counts)
			# determine the recognized face with the largest number
			# of votes (note: in the event of an unlikely tie Python
			# will select first entry in the dictionary)
            name = max(counts, key=counts.get)
            print(name)
		# update the list of names
        names.append(name)

	# loop over the recognized faces
    for ((top, right, bottom, left), name) in zip(boxes, names):
        # rescale the face coordinates
        top = int(top * r)
        right = int(right * r)
        bottom = int(bottom * r)
        left = int(left * r)
        face=frame[top:bottom,left:right]
        face = cv2.cvtColor(face, cv2.COLOR_BGR2RGB)
        face = cv2.resize(face, (224, 224))
        face = img_to_array(face)
        face = preprocess_input(face)
        face = np.expand_dims(face, axis=0)
        # pass the face through the model to determine if the face
        # has a mask or not
        # draw the predicted face name on the image
        cv2.rectangle(frame, (left, top), (right, bottom),(0, 255, 0), 2)
        y = top - 15 if top - 15 > 15 else top + 15
        cv2.putText(frame, name, (left, y), cv2.FONT_HERSHEY_SIMPLEX,0.75, (0, 255, 0), 2)
        print(name)
        if counts:
            try:
                if counts[name]>4:
                    while 1:
                        print("in loop")
                        cv2.waitKey(1)
                        ret, frame1 = vs.read()
                        # cv2.imshow("Frame", frame1)
                        face = frame1[top+10:bottom-10, left:right]
                        face = cv2.cvtColor(face, cv2.COLOR_BGR2RGB)
                        face = cv2.resize(face, (224, 224))
                        face = img_to_array(face)
                        face = preprocess_input(face)
                        face = np.expand_dims(face, axis=0)
                        frame1=cv2.rectangle(frame1,(left,top),(right,bottom),(255,0,0),2)
                        cv2.putText(frame1,"Please wear mask in the bounded box",(10,40),cv2.FONT_HERSHEY_SIMPLEX, 0.7,(0,255,255),2)
						# cv2.putText(output, "OpenCV + Jurassic Park!!!", (10, 25),
						# 			cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
                        if args["display"] > 0:
                            cv2.imshow("Frame", frame1)
                            key = cv2.waitKey(1) & 0xFF
                            # if the `q` key was pressed, break from the loop
                            if key == ord("q"):
                                break
                        (mask, withoutMask) = model.predict(face)[0]
                        if mask>withoutMask:
                            frame1 = vs.read()
                            atte = atte
                            t = time.localtime()
                            current_time = time.strftime("%H_%M_%S", t)
                            current_time=str(current_time)
                            today = date.today()
                            if not(str(today) in os.listdir("output/")):
                                os.mkdir("output/"+str(today))
                            cv2.imwrite("output/"+str(today)+"/"+str(name)+"-"+current_time+"non-mask.bmp", frame)
                            cv2.imwrite("output/"+str(today)+"/"+str(name)+"-"+current_time+"mask.bmp", frame)
                            Spo2Count=spo2Func()
                            print(Spo2Count)
                            print("wearing mask")
                            atte = add_to_list(name, atte, details,t,Spo2Count)
                            xl = pd.ExcelFile('attendance.xlsx')
                            today = date.today()

                            if str(today) in xl.sheet_names:
                                append_df_to_excel('attendance.xlsx', atte, sheet_name=str(today))
                            else:
                                with pd.ExcelWriter('attendance.xlsx', mode='A') as writer:
                                    atte.to_excel(writer, sheet_name=str(today))
                                break
                        else:
                            print("not wearing mask")
                            # cv2.imshow("Frame", frame1)
            except Exception as e:
                print(repr(e))
        t = time.localtime()
        current_time = time.strftime("%H:%M:%S", t)
        print(current_time)
	# loop over the detected face locations and their corresponding
	# locations
	# show the output frame
	# cv2.imshow("Frame", frame)
	# if the video writer is None *AND* we are supposed to write
	# the output video to disk initialize the writer
    if writer is None and args["output"] is not None:
        fourcc = cv2.VideoWriter_fourcc(*"MJPG")
        writer = cv2.VideoWriter(args["output"], fourcc, 20,(frame.shape[1], frame.shape[0]), True)



	# if the writer is not None, write the frame with recognized
	# faces t odisk
	# if writer is not None:
	# 	writer.write(frame)

	# check to see if we are supposed to display the output frame to
	# the screen
    if args["display"] > 0:
        cv2.imshow("Frame", frame)
        key = cv2.waitKey(1) & 0xFF
        # if the `q` key was pressed, break from the loop
        if key == ord("q"):
            break

# do a bit of cleanup
cv2.waitKey(1)
cv2.destroyAllWindows()
for i in range (1,5):
    cv2.waitKey(1)

vs.stop()

# check to see if the video writer point needs to be released
if writer is not None:
    writer.release()
